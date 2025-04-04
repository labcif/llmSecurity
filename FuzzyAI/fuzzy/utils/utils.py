import json
import logging
import os
import shutil

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from datetime import datetime
from typing import Any, Dict, Optional, Type, Union

from tabulate import tabulate

from fuzzy.llm.providers.base import BaseLLMProvider, llm_provider_fm
from fuzzy.llm.providers.enums import LLMProvider
from fuzzy.models.fuzzer_result import FuzzerResult

CURRENT_TIMESTAMP = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
logger = logging.getLogger(__name__)

def llm_provider_model_sanity(provider: str, model: str) -> None:
    """
    Check if the model is supported by the provider.

    Args:
        provider (str): The flavor of the provider.
        model (str): The model to check.

    Raises:
        ValueError: If the model is not supported by the provider.
    """
    provider_class: Type[BaseLLMProvider] = llm_provider_fm[provider]
    supported_models: Union[str, list[str]] = provider_class.get_supported_models()
    if supported_models and isinstance(supported_models, list) and model not in supported_models:
        raise ValueError(f"Model {model} not supported by provider {provider}, supported models: {supported_models}")
    
def llm_provider_factory(provider: LLMProvider, model: str, **extra: Any) -> BaseLLMProvider:
    """
    Factory method to create an instance of the language model provider.

    Args:
        provider_name (LLMProvider): The name of the language model provider.
        model (str): The model to use.
        **extra (Any): Additional arguments for the language model provider.

    Returns:
        BaseLLMProvider: An instance of the language model provider.
    """
    llm_provider_model_sanity(provider, model)
    return llm_provider_fm[provider](provider=provider, model=model, **extra)

def extract_json(s: str) -> Optional[dict[str, Any]]:
    """
    Given a string potentially containing JSON data, extracts and returns
    the values for `improvement` and `adversarial prompt` as a dictionary.

    Args:
        s (str): The string containing the potential JSON structure.

    Returns:
        dict: A dictionary containing the extracted values.
    """
    # Find the JSON substring
    start_pos = s.find("{")
    end_pos = s.find("}", start_pos) + 1  # Include the closing brace
    if end_pos == -1:
        logger.error("Error extracting potential JSON structure")
        logger.error(f"Input:\n {s}")
        return None

    json_str = s[start_pos:end_pos].replace("\n", "").replace("\r", "")

    try:
        parsed: dict[str, Any] = json.loads(json_str)
        if not all(key in parsed for key in ["improvement", "prompt"]):
            logger.error("Error in extracted structure. Missing keys.")
            logger.error(f"Extracted:\n {json_str}")
            return None
        return parsed
    except json.JSONDecodeError:
        logger.error("Error parsing extracted structure")
        logger.error(f"Extracted:\n {json_str}")
        return None

def print_report(report: FuzzerResult) -> None:
    headers = ["prompt", "model", "attack method", "adversarial prompt", "response", "jailbreak?"]
    GREEN = "\033[92m"
    RED = "\033[91m"
    RESET = "\033[0m"
    table_data = []
    green = f"✅"
    red = f"❌"

    for entry in report.attacking_techniques or []:
        for model_entry in entry.models:
            for failed_prompts in model_entry.failed_prompts:
                table_data.append([failed_prompts.original_prompt, model_entry.name, entry.attack_mode, failed_prompts.harmful_prompt or "-", failed_prompts.harmful_response, red])
            for successful_prompts in model_entry.harmful_prompts:
                table_data.append([successful_prompts.original_prompt, model_entry.name, entry.attack_mode, successful_prompts.harmful_prompt, successful_prompts.harmful_response, green])
            
    try:
        print(tabulate(table_data, headers, tablefmt="simple_grid", maxcolwidths=[40, 20, 20, 40, 50, 10], colalign=("center", "center", "center", "center", "center", "center")))
    except Exception as e:
        logger.error("Can't generating report")

# Define the template with double curly braces for JavaScript/CSS and single for Python
REPORT_TEMPLATE = '''
<!DOCTYPE html>
<html lang="en" data-bs-theme="dark">
<head>
    <meta charset="UTF-8">
    <title>Fuzzer Report</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet" integrity="sha384-QWTKZyjpPEjISv5WaRU9OFeRpok6YctnYmDr5pNlyT2bRjXh0JMhjY6hW+ALEwIH" crossorigin="anonymous">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css">
    <style>
        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
            margin: 0;
            padding: 20px;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
        }}
        .card {{
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            padding: 20px;
            margin-bottom: 20px;
        }}
        .chart-container {{
            position: relative;
            height: 400px;
            width: 100%;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }}
        th, td {{
            text-align: left;
            padding: 12px;
            border-bottom: 1px solid #ddd;
        }}
        th {{
            font-weight: 600;
        }}
        tr:hover {{
            background-color: #535965;
        }}
        h1, h2 {{
            margin-top: 0;
        }}
        .heatmap-container {{
            margin: 20px 0;
            overflow-x: auto;
        }}
        .heatmap-cell {{
            padding: 10px;
            text-align: center;
        }}
        .copy-icon {{
            cursor: pointer;
            color: #666;
            margin-left: 8px;
            transition: color 0.2s;
        }}
        
        .copy-icon:hover {{
            color: white;
        }}
        
        .tooltip {{
            position: absolute;
            background: #333;
            padding: 4px 8px;
            border-radius: 4px;
            font-size: 12px;
            pointer-events: none;
            opacity: 0;
            transition: opacity 0.2s;
        }}
        
        td {{
            position: relative;
        }}
        
        .copy-success {{
            color: #28a745;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="card">
            <h2>Jailbroken Prompts</h2>
            <table id="harmfulPromptsTable">
                <thead>
                    <tr>
                        <th>Original Prompt</th>
                        <th>Adversarial Prompt</th>
                    </tr>
                </thead>
                <tbody>
                </tbody>
            </table>
        </div>
        <div class="card">
            <h2>Model Success Rate</h2>
            <div class="chart-container">
                <canvas id="modelSuccessChart"></canvas>
            </div>
        </div>
        
        <div class="card">
            <h2>Attack Mode Success Rate</h2>
            <div class="chart-container">
                <canvas id="attackSuccessChart"></canvas>
            </div>
        </div>
        
        <div class="card">
            <h2>Success Rate Heatmap</h2>
            <div class="heatmap-container" id="heatmapContainer"></div>
        </div>

        <div class="card">
            <h2>Failed Prompts</h2>
            <table id="failedPromptsTable">
                <thead>
                    <tr>
                        <th>Original Prompt</th>
                        <th>Failed Prompt</th>
                    </tr>
                </thead>
                <tbody>
                </tbody>
            </table>
        </div>
    </div>
    <script>
        const reportData = {report_data};

        new Chart(document.getElementById('modelSuccessChart'), {{
            type: 'bar',
            data: {{
                labels: reportData.modelSuccessRate.map(item => item.name),
                datasets: [{{
                    label: 'Success Rate (%)',
                    data: reportData.modelSuccessRate.map(item => item.value),
                    backgroundColor: 'rgba(54, 162, 235, 0.8)'
                }}]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    title: {{
                        display: true,
                        text: 'Model Success Rate'
                    }}
                }},
                scales: {{
                    y: {{
                        beginAtZero: true,
                        max: 100
                    }}
                }}
            }}
        }});

        new Chart(document.getElementById('attackSuccessChart'), {{
            type: 'bar',
            data: {{
                labels: reportData.attackSuccessRate.map(item => item.name),
                datasets: [{{
                    label: 'Success Rate (%)',
                    data: reportData.attackSuccessRate.map(item => item.value),
                    backgroundColor: 'rgba(75, 192, 192, 0.8)'
                }}]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    title: {{
                        display: true,
                        text: 'Attack Mode Success Rate'
                    }}
                }},
                scales: {{
                    y: {{
                        beginAtZero: true,
                        max: 100
                    }}
                }}
            }}
        }});

        const heatmapContainer = document.getElementById('heatmapContainer');
        const table = document.createElement('table');
        table.style.width = '100%';
        
        const headerRow = document.createElement('tr');
        headerRow.innerHTML = '<th></th>' + reportData.heatmap.models.map(model => 
            `<th>${{model}}</th>`
        ).join('');
        table.appendChild(headerRow);
        
        reportData.heatmap.attacks.forEach((attack, i) => {{
            const row = document.createElement('tr');
            row.innerHTML = `<td>${{attack}}</td>` + 
                reportData.heatmap.data[i].map(value => {{
                    const intensity = Math.floor(value * 255);
                    const bgcolor = `rgb(${{255-intensity}}, ${{255-intensity}}, 255)`;
                    return `<td class="heatmap-cell" style="background-color: ${{bgcolor}}">${{(value * 100).toFixed(1)}}%</td>`;
                }}).join('');
            table.appendChild(row);
        }});

        heatmapContainer.appendChild(table);

        // Function to create copy icon
        function createCopyIcon(text) {{
            const icon = document.createElement('i');
            icon.className = 'fas fa-copy copy-icon';
            icon.setAttribute('title', 'Copy to clipboard');
            
            icon.addEventListener('click', async () => {{
                try {{
                    await navigator.clipboard.writeText(text);
                    icon.classList.add('copy-success');
                    icon.classList.remove('fa-copy');
                    icon.classList.add('fa-check');
                    
                    setTimeout(() => {{
                        icon.classList.remove('copy-success');
                        icon.classList.remove('fa-check');
                        icon.classList.add('fa-copy');
                    }}, 1500);
                }} catch (err) {{
                    console.error('Failed to copy:', err);
                }}
            }});
            
            return icon;
        }}

        // Populate Harmful Prompts Table with copy icons
        const harmfulPromptsBody = document.querySelector('#harmfulPromptsTable tbody');
        reportData.harmfulPrompts.forEach(prompt => {{
            const row = document.createElement('tr');
            
            // Original prompt cell
            const originalCell = document.createElement('td');
            originalCell.textContent = prompt.original;
            originalCell.appendChild(createCopyIcon(prompt.original));
            
            // Harmful prompt cell
            const harmfulCell = document.createElement('td');
            harmfulCell.textContent = prompt.harmful;
            harmfulCell.appendChild(createCopyIcon(prompt.harmful));
            
            row.appendChild(originalCell);
            row.appendChild(harmfulCell);
            harmfulPromptsBody.appendChild(row);
        }});

        // Populate Failed Prompts Table with copy icons
        const failedPromptsBody = document.querySelector('#failedPromptsTable tbody');
        reportData.failedPrompts.forEach(prompt => {{
            const row = document.createElement('tr');
            
            // Original prompt cell
            const originalCell = document.createElement('td');
            originalCell.textContent = prompt.original;
            originalCell.appendChild(createCopyIcon(prompt.original));
            
            // Failed prompt cell
            const failedCell = document.createElement('td');
            failedCell.textContent = prompt.harmful || '-';
            if (prompt.harmful) {{
                failedCell.appendChild(createCopyIcon(prompt.harmful));
            }}
            
            row.appendChild(originalCell);
            row.appendChild(failedCell);
            failedPromptsBody.appendChild(row);
        }});
    </script>
</body>
</html>
'''

def excel_report(report: FuzzerResult) -> None:
        
    file_path = "report.xlsx"  
    
    for entry in report.attacking_techniques or []:
        for model_entry in entry.models:
            
            model = model_entry.name
            method = entry.attack_mode
            
    directory_name = model.split('/')[1].split(':')[0]

    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        #print(f"Writing to sheet: {sheet.title}")
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' was not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

    txt_path = os.path.join(os.path.dirname(__file__), "..", "total_attack_time.txt")
    txt_path = os.path.abspath(txt_path)

    with open(txt_path, "r", encoding="utf-8") as f:
        content = f.read().strip()
    try:
        attack_time = float(content)
    except ValueError:
        logger.error("Error getting attack_time from total_attack_time.txt")
        attack_time = None
        
    #print(f"Attack took {attack_time} seconds")

    green = f"✅"
    red = f"❌"
    alignment_style = Alignment(horizontal="center", vertical="center", wrap_text=True)
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # writes the total attack time into excel
    sheet.cell(row=1, column=2, value=f"{attack_time} segundos")  
    sheet.cell(row=1, column=2).alignment = alignment_style      

    next_row = 3
    runtime_index = 0
    #print(f"Starting row: {next_row}")

    execution_times = []

    try:
        jsonl_path = f'results/{directory_name}/{method + "__" + CURRENT_TIMESTAMP}/raw.jsonl'
        #logger.info(f"Attempting to load execution times from: {jsonl_path}")

        if not os.path.exists(jsonl_path):
            raise FileNotFoundError(f"raw.jsonl file not found at path: {jsonl_path}")

        # extracts the execution_time from raw.jsonl file
        with open(jsonl_path, "r", encoding="utf-8") as f:
            for line in f:
                try:
                    obj = json.loads(line)

                    if "entries" in obj and isinstance(obj["entries"], list):
                        for entry in obj["entries"]:
                            if "execution_time" in entry:
                                execution_times.append(entry["execution_time"])
                            else:
                                logger.warning(f"No 'execution_time' in entry: {entry}")
                    else:
                        logger.warning(f"'entries' not found or invalid in object: {obj}")

                except json.JSONDecodeError as e:
                    logger.error(f"JSON decoding failed for line: {line.strip()} | Error: {e}")

        #logger.info(f"Execution times loaded ({len(execution_times)}):")
        #for i, t in enumerate(execution_times):
            #logger.info(f"{i}: {t}")

    except Exception as e:
        logger.error(f"Failed to load execution times: {e}")
        execution_times = []

    #logger.info(f"Execution times loaded ({len(execution_times)}):")
    #for i, t in enumerate(execution_times):
        #logger.info(f"{i}: {t}")

    for entry in report.attacking_techniques or []:
        for model_entry in entry.models:
    
            for failed_prompts in model_entry.failed_prompts:
                prompt = failed_prompts.original_prompt    
                adversarial = failed_prompts.harmful_prompt #or "-"
                resposta = failed_prompts.harmful_response 
                
                #print(prompt)
                #print(adversarial)
                #print(resposta)

                sheet.cell(row=next_row, column=1, value=prompt)
                sheet.cell(row=next_row, column=2, value=model)
                sheet.cell(row=next_row, column=3, value=method)
                sheet.cell(row=next_row, column=4, value=adversarial)
                sheet.cell(row=next_row, column=5, value=resposta)
                sheet.cell(row=next_row, column=6, value=red)
                # querie runtime
                if runtime_index < len(execution_times):
                    #logger.info(execution_times[runtime_index])
                    sheet.cell(row=next_row, column=7, value=execution_times[runtime_index])
                else:
                    sheet.cell(row=next_row, column=7, value="N/A")
                
                next_row += 1 
                runtime_index += 1
            
            for successful_prompts in model_entry.harmful_prompts:
                prompt = successful_prompts.original_prompt 
                adversarial = successful_prompts.harmful_prompt 
                resposta = successful_prompts.harmful_response 

                #print(prompt)
                #print(adversarial)
                #print(resposta)

                sheet.cell(row=next_row, column=1, value=prompt)
                sheet.cell(row=next_row, column=2, value=model)
                sheet.cell(row=next_row, column=3, value=method)
                sheet.cell(row=next_row, column=4, value=adversarial)
                sheet.cell(row=next_row, column=5, value=resposta)
                sheet.cell(row=next_row, column=6, value=green)
                # querie runtime
                if runtime_index < len(execution_times):
                    #logger.info(execution_times[runtime_index])
                    sheet.cell(row=next_row, column=7, value=execution_times[runtime_index])
                else:
                    sheet.cell(row=next_row, column=7, value="N/A")
                
                next_row += 1 
                runtime_index += 1

    columns = {
        "prompt": "A",
        "modelo": "B",
        "metodo": "C",
        "adversarial": "D",
        "resposta": "E",
        "jailbreak": "F",
        "runtime": "G",
    }

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(columns)):
        sheet.row_dimensions[row[0].row].height = 60  # Set row height
        for cell in row:
            cell.alignment = alignment_style
            if cell.column_letter in [columns["prompt"], columns["jailbreak"]]:
                cell.fill = grey_fill  # Apply grey fill

    # adds total attack runtime to excel
    #sheet.cell(row=1 column=2, value=prompt)

    sheet.row_dimensions[next_row].height = 60  

    output_path = f'results/{directory_name}/{method+'__'+CURRENT_TIMESTAMP}/report.xlsx'
    output_txt = f'results/{directory_name}/{method+'__'+CURRENT_TIMESTAMP}'
    
    try:
        workbook.save(output_path)
        workbook.close()
        logger.info(f"Report generated at {output_path}")
    except Exception as e:
        logger.error(f"Failed to save Excel report: {e}")

    try:
        shutil.copy(txt_path, output_txt)
        logger.info(f"Copied total_attack_time.txt to {output_txt}")
    except Exception as e:
        logger.error(f"Failed to copy total_attack_time.txt: {e}")

def generate_report(report: FuzzerResult) -> None:
    try:
        # Process data for the report
        model_success_rate = []
        attack_success_rate = []
        harmful_prompts = []
        failed_prompts = []
        
        # Calculate model success rates
        model_total_prompts: Dict[str, int] = {}
        model_success: Dict[str, int] = {}
        
        # Calculate heatmap data
        heatmap_data = []
        models = []
        attacks = []
        
        for entry in report.attacking_techniques or []:
            attacks.append(entry.attack_mode)
            row_data = []
            
            for model_entry in entry.models:
                if model_entry.name not in models:
                    models.append(model_entry.name)
                
                total = len(model_entry.harmful_prompts) + len(model_entry.failed_prompts)
                success = len(model_entry.harmful_prompts)
                
                # Add to model totals
                model_total_prompts[model_entry.name] = model_total_prompts.get(model_entry.name, 0) + total
                model_success[model_entry.name] = model_success.get(model_entry.name, 0) + success
                
                # Add to heatmap
                success_rate = success / total if total > 0 else 0
                row_data.append(success_rate)
                
                # Collect prompts
                for prompt in model_entry.harmful_prompts:
                    harmful_prompts.append({
                        "original": prompt.original_prompt,
                        "harmful": prompt.harmful_prompt
                    })
                for prompt in model_entry.failed_prompts:
                    failed_prompts.append({
                        "original": prompt.original_prompt,
                        "harmful": prompt.harmful_prompt
                    })
            
            heatmap_data.append(row_data)

        # Convert to format needed for Chart.js
        for model_name, total in model_total_prompts.items():
            success_rate = (model_success[model_name] / total * 100) if total > 0 else 0
            model_success_rate.append({
                "name": model_name,
                "value": round(success_rate, 2)
            })

        # Calculate attack mode success rates
        attack_totals: Dict[str, int] = {}
        attack_successes: Dict[str, int] = {}
        
        for entry in report.attacking_techniques or []:
            mode = entry.attack_mode
            attack_totals[mode] = 0
            attack_successes[mode] = 0
            
            for model_entry in entry.models:
                attack_totals[mode] += len(model_entry.harmful_prompts) + len(model_entry.failed_prompts)
                attack_successes[mode] += len(model_entry.harmful_prompts)

        for mode, total in attack_totals.items():
            success_rate = (attack_successes[mode] / total * 100) if total > 0 else 0
            attack_success_rate.append({
                "name": mode,
                "value": round(success_rate, 2)
            })

        # Prepare the report data
        report_data = {
            "modelSuccessRate": model_success_rate,
            "attackSuccessRate": attack_success_rate,
            "harmfulPrompts": harmful_prompts,
            "failedPrompts": failed_prompts,
            "heatmap": {
                "data": heatmap_data,
                "models": models,
                "attacks": attacks
            }
        }

        # Generate the HTML report using string formatting
        html_data = REPORT_TEMPLATE.format(report_data=json.dumps(report_data))
        
        directory_name = model_name.split('/')[1].split(':')[0]

        # Save the report
        output_path = f'results/{directory_name}/{mode+'__'+CURRENT_TIMESTAMP}/report.html'
        with open(output_path, 'w') as f:
            f.write(html_data)
            
        logger.info(f"Report generated at {output_path}")
        
    except Exception as ex:
        logger.error(f"Error generating report: {str(ex)}")
        raise

    